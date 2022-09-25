from ast import Break
import os
import datetime
from base64 import urlsafe_b64decode, urlsafe_b64encode
import filecmp
from ftplib import MSG_OOB
from io import SEEK_END
import json
from lib2to3.pytree import LeafPattern
from logging import INFO
from pydoc import describe
from random import random
from re import T
import re
import sqlite3
from time import time
from tracemalloc import stop
from turtle import st
from winsound import PlaySound
from bs4 import BeautifulSoup
import discord
import asyncio
from discord import player  
from discord.channel import VoiceChannel
from discord.player import FFmpegPCMAudio, PCMAudio
from idna import valid_contextj
import openpyxl
from json import loads
import bs4
import requests
from youtube_dl import YoutubeDL
import datetime
import bs4 
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from discord.utils import get
from discord import FFmpegPCMAudio
import asyncio
from random import *
from random import shuffle
import random
from random import random
from random import shuffle
import openpyxl
from openpyxl import *
from openpyxl import load_workbook, Workbook
import requests
from requests import *
import sqlite3
from youtube_dl import YoutubeDL
import bs4
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from discord.utils import get
from discord import FFmpegPCMAudio
import asyncio
import time
import openpyxl
from discord import commands
from discord_buttons_plugin import  *
import requests



jinbot = "진봇"
version = "1.9.1"
upa = datetime.datetime.today().year
upb = datetime.datetime.today().month
upc = datetime.datetime.today().day
upd = datetime.datetime.today().hour
upe = datetime.datetime.today().minute
upf = datetime.datetime.today().second
client = discord.Client()
@client.event
async def on_ready():
    print(client.user.name)
    print("성공적으로 봇이 시작되었습니다.")
    starttrue = 1
    game = discord.Game('도움말:진봇 도움말 문의하기:Jin_3070#7131 Jinbot ' + version + 'Ver!')
    await client.change_presence(status=discord.Status.online, activity=game)
word = ["가게른", "가녘", "가돌리늄", "가랒", "가벼운피읖", "가재무릊", "가짓부렁", "가쿄", "갈루아몽브룅", "갈이그릇", "강녘", "개노릇", ]

start = 1

bot = commands.Bot(command_prefix="진봇 ")

@client.event 
async def on_message(message):
    if message.content == "어어":
        await message.channel.send("어엉")

    if message.content == "뿗쀓삙":
        i = 1
        while True:
            if i == 20:
                break
            else:
                i = i =+ 1
                await message.channel.send(f"{message.author.mention}")
    if message.content == "진봇 가위바위보 바위":
        rsp = int(random() * 3)
        if rsp == 1:
            embed = discord.Embed(title = "이겼닼ㅋㅋㅋ", description = "가위바위보 못하누 ㅋㅋ", color=0x00ff00)
            await message.channel.send(embed=embed)
        elif rsp == 2:
            embed = discord.Embed(title = "응 비김", description = "넌 날 못이겨")
            await message.channel.send(embed=embed)
        elif rsp == 3:
            embed = discord.Embed(title = "아 졌네", description = "어 철가위니까 돌을 깡깡 내려치면 언젠간 돌이 먼저 부서지죠?(억지)", color=0xff0000)
            await message.channel.send(embed=embed)
    if message.content == "진봇 가위바위보 가위":
        rsp = int(random() * 3)
        if rsp == 1:
            embed = discord.Embed(title = "이겼닼ㅋㅋㅋ", description = "가위바위보 못하누 ㅋㅋ", color=0x00ff00)
            await message.channel.send(embed=embed)
        elif rsp == 2:
            embed = discord.Embed(title = "응 비김", description = "넌 날 못이겨")
            await message.channel.send(embed=embed)
        elif rsp == 3:
            embed = discord.Embed(title = "아 졌네", description = "니가 살짝 늦게냄(?)이판 무효임 숫1고", color=0xff0000)
            await message.channel.send(embed=embed)
    if message.content == "진봇 가위바위보 보자기":
        rsp = int(random() * 3)
        if rsp == 1:
            embed = discord.Embed(title = "이겼닼ㅋㅋㅋ", description = "가위바위보 못하누 ㅋㅋ", color=0x00ff00)
            await message.channel.send(embed=embed)
        elif rsp == 2:
            embed = discord.Embed(title = "응 비김", description = "넌 날 못이겨")
            await message.channel.send(embed=embed)
        elif rsp == 3:
            embed = discord.Embed(title = "아 졌네", description = "돌로 종이를 확 찢어버려!", color=0xff0000)
            await message.channel.send(embed=embed)
    if message.content == "진봇 쀍":
        await message.channel.send("뭐")
    if message.content == "진봇 뷁":
        await message.channel.sned("뭐!!!")
    if message.content == "진봇 좋아?":
        await message.channel.send("뭔진 모르겠지만 일단 싫을듯")
    if message.content == "wlsqht":
        await message.channel.send("한영키나 바꾸세요")
    if message.content == "wㅣㄴ봇":
        await message.channel.send("디코 버-그")
    if message.content == ("진봇 핑"):
        embed = discord.Embed(title = ':ping_pong: 퐁!', description = str(client.latency) + 'ms', color = 0x00ff00)
        await message.channel.send(embed=embed)
    if message.content == "진봇 놀아줘":
        playsimsim = int(random() * 90)
    
        if playsimsim == 2:
            await message.channel.send("서울에서 생산된 고기는?\n||살코기(지방이 없으니까)||")
        if playsimsim == 1:
            await message.channel.send("지구가 제일 힘들었던 시기는?\n||고생대||")
        elif playsimsim == 3:
            await message.channel.send("F는 좀 애매한 이유는?\n||F = ma니까||")
        elif playsimsim == 4:
            await message.channel.send("me too와 비슷한 말은?\n||another(어 나도)||")
        elif playsimsim == 5:
            await message.channel.send("메이저리그에서 던지는 공은?\n||해외직구||")
        elif playsimsim == 6:
            await message.channel.send("이야기를 팔아서 돈 버는 사람은?\n||썰(Ssul)매장 주인||")
        elif playsimsim == 7:
            await message.channel.send("할아버지를 영어로?\n||맨들||")
        elif playsimsim == 8:
            await message.channel.send("옷감을 울게 만드는 것은?\n||섬유유(ㅠㅠ)연제||")
        elif playsimsim == 9:
            await message.channel.send("비빔밥 재료가 맨날 남는 이유는?\n||나물 음식이라서||")
        elif playsimsim == 10:
            await message.channel.send("도서관에서 음료수를 원샷하는 이유는?\n||반입금지라서||")
        elif playsimsim == 11:
            await message.channel.send("피카츄가 담배를 피기 전에 하는 말은?\n||피까||")
        elif playsimsim == 12:
            await message.channel.send("모든 사람을 일어나게 하는 숫자는?\n||다섯||")
        elif playsimsim == 13:
            await message.channel.send("호랑이(사자)가 항상 숙제를 안 하는 이유는?\n||밀림의 제왕이라서||")
        elif playsimsim == 14:
            await message.channel.send("자동차를 톡하고 치면?\n||카톡||")
        elif playsimsim == 15:
            await message.channel.send("다리미가 좋아하는 음식은?\n||피자||")
        elif playsimsim == 16:
            await message.channel.send("제일 폭발하기 쉬운 나라는?\n||부탄||")
        elif playsimsim == 17:
            await message.channel.send("뽑으면 우는 식물은?\n||우엉||")
        elif playsimsim == 18:
            await message.channel.send("광부가 가장 많은 나라는?\n||케냐||")
        elif playsimsim == 19:
            await message.channel.send("사람이 울면?\n||흑인||")
        elif playsimsim == 20:
            await message.channel.send("오키나와에 가면 살 찌는 이유는?\n||5끼 나오니까||")
        elif playsimsim == 21:
            await message.channel.send("소나무가 삐지면?\n||칫솔||")
        elif playsimsim == 22:
            await message.channel.send("우유가 아프면?\n||앙팡||")
        elif playsimsim == 23:
            await message.channel.send("등쳐먹고 사는 사람은?\n||안마사||")
        elif playsimsim == 24:
            await message.channel.send("소가 불에 타면?\n||탄소||")
        elif playsimsim == 25:
            await message.channel.send("오리가 얼면?\n||언덕||")
        elif playsimsim == 26:
            await message.channel.send("가장 비싼 새는?\n||백조||")
        elif playsimsim == 27:
            await message.channel.send("땅은 어떻게 울지?\n||흙흙||")
        elif playsimsim == 28:
            await message.channel.send("수박 한 통에 1만 원인데 두통엔?\n||게보린||")
        elif playsimsim == 29:
            await message.channel.send("맥주가 죽기 전에 남긴 말은?\n||유언비어||")
        elif playsimsim ==30:
            await message.channel.send("세상에서 가장 가난한 왕은?\n||최저임금||")
        elif playsimsim == 31:
            await message.channel.send("칼이 정색하면?\n||검정색||")
        elif playsimsim == 32:
            await message.channel.send("가수 설운도가 옷을 벗는 순서가 어떻게 될까?\n||상하의||")
        elif playsimsim == 33:
            await message.channel.send("항상 미안해하는 연예인은?\n||지성||")
        elif playsimsim == 34:
            await message.channel.send("반성문을 영어로 하면?\n||글로벌||")
        elif playsimsim == 35:
            await message.channel.send("세종대왕이 나온 고등학교는?\n||가갸거겨고교||")
        elif playsimsim == 36:
            await message.channel.send("새우가 주인공인 드라마는?\n||대하드라마||")
        elif playsimsim == 37:
            await message.channel.send("화장실에서 금방 나온 사람은?\n||일본사람||")
        elif playsimsim == 38:
            await message.channel.send("회를 가장 잘 뜨는 곳은?\n||||")
        elif playsimsim == 39:
            await message.channel.send("소금의 유통기한은?\n||천일염||")
        elif playsimsim == 40:
            await message.channel.send("왕이 궁에 가기 싫은 때 하는 말은?\n||궁시렁궁시렁||")
        elif playsimsim == 41:
            await message.channel.send("아마존에 살고 있는 사람의 이름은?\n||아마 존?||")
        elif playsimsim == 42:
            await message.channel.send("신발이 화가 나면?\n||신발 끈||")
        elif playsimsim == 43:
            await message.channel.send("고등학생이 싫어하는 나무는?\n||야자나무||")
        elif playsimsim == 44:
            await message.channel.send("세상에서 가장 억울한 도형은?\n||원통||")
        elif playsimsim == 45:
            await message.channel.send("세상에서 가장 지루한 중학교?\n||로딩중||")
        elif playsimsim == 46:
            await message.channel.send("서울이 추우면?\n||서울시립대||")
        elif playsimsim == 47:
            await message.channel.send("정말 큰 학은?\n||대학||")
        elif playsimsim == 48:
            await message.channel.send("가장 야한 가수는?\n||다비치||")
        elif playsimsim == 49:
            await message.channel.send("어부들이 싫어하는 사람은?\n||배철수||")
        elif playsimsim == 50:
            await message.channel.send("얼음이 죽으면?\n||다이빙||")
        elif playsimsim == 51:
            await message.channel.send("이탈리아의 날씨?\n||습하겟디||")
        elif playsimsim == 52:
            await message.channel.send("세상에서 가장 뜨거운 바다는?\n||열받아||")
        elif playsimsim == 53:
            await message.channel.send("인천 앞바다의 반대는?\n||인천 엄마다||")
        elif playsimsim == 54:
            await message.channel.send("누룽지를 영어로 하면?\n||바비브라운||")
        elif playsimsim == 55:
            await message.channel.send("가는 말이 고우면 오는 말은?\n||Come||")
        elif playsimsim == 56:
            await message.channel.send("높은 곳에서 새끼를 낳는 동물은?\n||하이에나||")
        elif playsimsim == 57:
            await message.channel.send("딸기가 도망가면?\n||딸기잼||")
        elif playsimsim == 58:
            await message.channel.send("길 가다 음식을 주운 것을 3글자로?\n||푸드 득||")
        elif playsimsim == 59:
            await message.channel.send("옥수수가 시험을 보는 것을 4글자로?\n||콘테스트||")
        elif playsimsim == 60:
            await message.channel.send("칼로 산을 두 개로 자르면?\n||두산베어스||")
        elif playsimsim == 61:
            await message.channel.send("낙옆이 애교를 부리면서 한 말?\n||나기엽?||")
        elif playsimsim == 62:
            await message.channel.send("신부가 부케를 던지면서 하는 말\n||기다려 본캐로 온다||")
        elif playsimsim == 63:
            await message.channel.send("물리치료를 받는 이유는?\n||병을 물리치료고||")
        elif playsimsim == 64:
            await message.channel.send("부침개를 먹으면 살이 찌는 이유는?\n||전 부처먹기 때문||")
        elif playsimsim == 65:
            await message.channel.send("우유가 넘어지면?\n||아야||")
        elif playsimsim == 66:
            await message.channel.send("보내기 싫으면?\n||가위나 바위를 낸다.||")
        elif playsimsim == 67:
            await message.channel.send("세상에서 가장 놀라운 거품은?\n||언빌리버블||")
        elif playsimsim == 68:
            await message.channel.send("옳지 못한 오리는?\n||부도덕||")
        elif playsimsim == 69:
            await message.channel.send("모내기 기계가 고장이 나면?\n||심기 불편||")
        elif playsimsim == 70:
            await message.channel.send("흑형의 여자친구는?\n||깜짝||")
        elif playsimsim == 71:
            await message.channel.send("간첩이 동물원에 가면 찾는 새는?\n||공작새||")
        elif playsimsim == 72:
            await message.channel.send("오래된 우유를 다른 말로?\n||연세우유||")
        elif playsimsim == 73:
            await message.channel.send("은행 금고를 1원도 남기지 않고 털면?\n||클린턴||")
        elif playsimsim == 74:
            await message.channel.send("모태솔로를 영어로 하면?\n||없데이트||")
        elif playsimsim == 75:
            await message.channel.send("애플이 망한 이유는?\n||MAC(막) 만들어서||")
        elif playsimsim == 76:
            await message.channel.send("돼지가 금값인 이유?\n||돼지가 '한 돈' 가격이라서||")
        elif playsimsim == 77:
            await message.channel.send("연필로 잘못 썼을 때 피카츄로 지우면 되는 이유는?\n||지우 개니까||")
        elif playsimsim == 78:
            await message.channel.send("아이언맨을 대체할 히어로가 없는 이유는?\n||iron 사람 또 없어서||")
        elif playsimsim == 79:
            await message.channel.send("흑인이 도망갈 때 하는 말?\n||니거기서!||")
        elif playsimsim == 80:
            await message.channel.send("해산물이 전부 '부패'하는 이유는?\n||생선은 전부 비리니까||")

        elif playsimsim == 81:
            await message.channel.send(file=discord.File("1.gif"))
        elif playsimsim == 82:
            await message.channel.send(file=discord.File("2.gif"))
        elif playsimsim == 83:
            await message.channel.send(file=discord.File("3.gif"))

        elif playsimsim == 84:
            await message.channel.send(file=discord.File("4.gif"))
        elif playsimsim == 85:
            await message.channel.send(file=discord.File("5.gif"))
        elif playsimsim == 86:
            await message.channel.send(file=discord.File("6.gif"))
        elif playsimsim == 87:
            await message.channel.send(file=discord.File("7.gif"))
        elif playsimsim == 88:
            await message.channel.send(file=discord.File("8.gif"))
        elif playsimsim == 89:
            await message.channel.send(file=discord.File("9.gif"))
        elif playsimsim == 90:
            await message.channel.send(file=discord.File("10.gif"))

        elif playsimsim == 91:
            await message.channel.send(file=discord.File("12.gif"))
        elif playsimsim == 92:
            await message.channel.send(file=discord.File("13.gif"))
        elif playsimsim == 93:
            await message.channel.send(file=discord.File("14.gif"))

            
    if message.content == "진봇 끝말잇기":
        @client.event
        async def on_message(message):
            if not message.content.startswith("진봇 끝말잇기"):
                stop
    if message.content == "진봇 베타":
        await message.channel.send(file=discord.File("1.gif"))
        await message.channel.send("현재 테스트중인 기능이에요! 곧 진봇 놀아줘 기능에 추가될거랍니다!")

    if message.content.startswith("진봇 킥"):
        if(message.author.guild_permissions.kick_members):
            user = message.content.split(" ")
            reason = message.content.split(" ")
            if reason[2] == None:
                reason[2] = "사유가 없습니다." 
                embed = discord.Embed(title="킥", description = user[1] + f"님은 {message.guild.name}서버에서 킥당했습니다. 사유는 다음과 같습니다.\n ```" + reason[2] + "```", color=0xff0000)
                await user[1].send(embed=embed)
            else:
                embed = discord.Embed(title="킥", description = user[1] + f"님은 {message.guild.name}서버에서 킥당했습니다. 사유는 다음과 같습니다.\n ```" + reason[2] + "```", color=0xff0000)
                await user[1].channel.send(embed=embed)
        else:
            await message.channel.send(embed=discord.Embed(title="권한 부족", description = message.author.mention + "님은 유저를 킥할 수 있는 권한이 없습니다.", color = 0xff0000))
            return



    if message.content == "진봇 업타임":
        await message.channel.send(str(upa) + "년 " + str(upb) + "월 " + str(upc) + "일 " + str(upd) + "시 " + str(upe) + "분 " + str(upf) + "초부터 작동중입니다.")
    if message.content == "진봇 ㅂㅇ":
        await message.channel.send("ㅂ2")
    if message.content == "진봇 빼액":
        await message.channel.send("시끄러")
    if message.content == "진봇 도움말":
        await message.channel.send("```진봇 뽑기:일정 확률로 당첨됩니다!\n진봇 서버목록:현재 진봇이 접속해있는 서버를 모두 보여줍니다!\n진봇 팀나누기:팀을 나눠줍니다!\n(예시:진봇 팀나누기 사람1 사람2 사람3 사람4/팀1 팀1 팀2 팀2)\n진봇 현재시각:현재시각을 보여줍니다!\n진봇 정보:자신의 닉네임, 서버닉네임, 가입일, ID를 보여줍니다!\n진봇 투표:투표합니다!(예시:진봇 투표 지금 기분 심심함/좋음/슬픔/나쁨/화남)\n진봇 가위바위보:가위바위보를 해드립니다!(진봇 가위바위보 보자기 or 바위 or 가위)\n진봇 음악(버그로 인해 잠시 중단)\n진봇 역사퀴즈:역사퀴즈를 내드립니다!\n진봇 타이머(버그로 인해 잠시 중단)\n진봇 놀아줘:아재개그를 던져드립니다!\n진봇 업타임:진봇이 언제부터 작동했는지 보여드립니다!\n진봇 따라해:따라합니다(진봇 따라해 진봇이 따라하면 좋겠는말)\n진봇 대화 도움말\n진봇 학습/기억:학습하거나 학습했던걸 기억해냅니다!(진봇 학습/진봇은/멍청이:앞으로 진봇 기억/진봇은 으로 할때 멍청이라고 대답합니다.)\n나머지는 사용하시면서 천천히 알아가주세요!(모든 명령어는 진봇 으로 시작한답니다!)\n아이디어 추천받습니다 아이디어 있으면 Jin_3070#7131로 알려줘요...```\n**Jinbot " + version + " Ver!**\nmade by Jin_3070#7131!" )
    if message.content == "진봇 뽑기":
        ran = int(random() * 8000000)
        await message.channel.send("이거 당첨되면 진이 문상준다는 소문이...\n참고:확률 1/9999999999999999임")
        if ran == 0:
            d = "당첨"
        else:
            d = "꽝"
        await message.channel.send(d)
    if message.content == "진봇 사과해":
        await message.channel.send("[저작권에 의해 삭제된 먹는사과 이미지입니다.]")
    if message.content == "진봇 복권":
        a = int(random() * 42)
        b = int(random() * 42)

    if message.content == "진봇":
        await message.channel.send("왜")
    if message.content == "진봇 개발자":
        await message.channel.send("Jin_3070#7131")
    if message.content == "진봇 버그":
        await message.channel.send("Jin_3070#7131로 dm ㄱㄱ")
    if message.content == "진봇 노잼":
        await message.channel.send("노잼이여서 어쩌라고")
    if message.content == "진봇 뭐해?":
        await message.channel.send("니랑 챝하고있잖아 그것도 모르냐")
    if message.content == "마요봇":
        await message.channel.send("마요도 맛없고 마요봇도 맛없어 글고 마요봇이랑 나랑 적대관계니까 알아둬")
    if message.content == "진봇 나가":
        await message.channel.send("NAGA")
    if message.content == "진봇 업데이트 내역":
        await message.channel.send("업데이트때 뭐했는지 기억안남 ㅅㄱ")
    if message.content == "진봇 스포일러":
        await message.channel.send("|||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||")
    #if message.content.startswith("진봇 청소"):
    #   number = (message.content.split("/")[5])
    #  await message.delete()
    # await message.channel.purge(limit=number)
        #await message.channel.send(f"{number}개의 메시지를 삭제하였습니다.")

    if message.content.startswith("진봇 타이머"):
        timerseting = int(message.content[6:].split("/"))
        await asyncio.sleep(timerseting)
        await message.channel.send(timerseting +"초가 지났습니다")

    if message.content == "진봇 맨션": 
        await message.channel.send(f"{message.author.mention} 옛다 맨션이다")
    if message.content == "진봇 버전":
        await message.channel.send("Jinbot " + version + "Ver!")

    


    if message.content.startswith("진봇 학습"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        learn = message.content.split("/")
        for i in range(1, 599999999999999999999):
            if sheet["A" + str(i)].value == "-" or None:
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("기억 완료!")
                break
        file.save("기억.xlsx")
    if message.content.startswith("진봇 기억"):
        file = openpyxl.load_workbook("기억.xlsx")
        sheet = file.active
        memory = message.content.split("/")
        for i in range(1, 50):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value)
                break

    if message.content == "진봇 가입":
        file = openpyxl.load_workbook("돈.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                await message.channel.send("이미 가입이 되어있습니다.")
                break
            elif sheet["A" + str(i)].value == "-" or None:
                sheet["A" + str(i)].value = str(message.author.id)
                sheet["B" + str(i)].value = "500"
                await message.channel.send("가입 완료!")
                break
        file.save("돈.xlsx")
    if message.content.startswith("진봇 도박"):
        file = openpyxl.load_workbook("돈.xlsx")
        sheet = file.active
        moneydobak = message.content.split("/")
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                if sheet["A" + str(i)].value == str(message.author.id):
                    ran = int(random() * 2) 
                    if ran == 0:
                        sheet["B" + str(i)].value = sheet["B" + str(i)].value + moneydobak
                        await message.channel.send("도박 성공!\n```+" + moneydobak + "```")
                        break
                    else:
                        sheet["B" + str(i)].value = sheet["B" + str(i)].value - moneydobak
                        await message.channel.send("도박 실패...\n```-" + moneydobak + "```")
                        break
        file.save("돈.xlsx")


    if message.content == "진봇 돈":
        file = openpyxl.load_workbook("돈.xlsx")
        i = 1
        sheet = file.active
        while True:
            if sheet["A" + str(i)].value == str(message.author.id):
                janak = sheet["B" + str(i)].value
                await message.channel.send("현재 잔액:" + janak)
                break
    
    


    if message.content == "진봇 ㄴㅇㄱ":
        await message.channel.send("ㄴㅇㄱ 뭐 뭐 어쩌라고")
    if message.content == "진봇 도배":
        await message.channel.send("도배하면 봇 이용정책 위반임 ㅋㅋㄹㅃㅃ")
    if message.content == "진봇 심심해":
        await message.channel.send("그럼 진봇 놀아줘를 입력해봐")
    if message.content == "진봇 잘가":
        await message.channel.send("잘간다는 개념이 뭐지 나는 움직일수가 없는데")
    if message.content == "진봇 ㅎㅇ":
        await message.channel.send("ㅎㅇ")
    if message.content == "진봇 안녕":
        await message.channel.send("ㅎㅇ")
    if message.content == "진봇 그냥":
        await message.channel.send("선생님특:숙제안했을때 그냥이라고하면 그냥이 어딨어라고함")
    if message.content == "진봇 어쩌라고":
        await message.channel.send("ㅇㅉㅌㅂ")
    if message.content == "진봇 ㅇㅉㄹㄱ":
        await message.channel.send("ㅇㅉㅌㅂ")
    if message.content == "진봇 ㅇㅉㄺ":
        await message.channel.send("ㅇㅉㄹㄱ임 ㅋㅋㅋ")
    if message.content == "진봇 ㄹㅇㅋㅋ":
        await message.channel.send("fdzz")
    if message.content == "진봇 ㅇㄱㄹㅇ":
        await message.channel.send("ㄹㅇㅋㅋ")
    if message.content == "진봇 ㅋ":
        await message.channel.send("아따 거짓말로 띠겁네~")
    if message.content == "진봇 ㅋㅋ":
        await message.channel.send("이거 보낸사람 무포정일듯")
    if message.content == "진봇 h":
        await message.channel.send("너는 나한테 h를 보내지만 나는 너에게 ㅗ를 보내노라")
    if message.content == "진봇 어몽어스":
        await message.channel.send("난 접었음")
    if message.content == "진봇 정체가 뭐야":
        await message.channel.send("||강제노역중 살려줘||")
    if message.content == "진봇 이스터에그":
        await message.channel.send("없음")
    if message.content == "진봇 데이터쪼가리":
        await message.channel.send("응 어쩔~")
    #if message.content.startswith(""):
        #file = openpyxl.load_workbook("레벨.xlsx")
        #sheet = file.active
        #exp = [10, 20, 30, 40, 50]
        #i = 1
        #while True:
            #if sheet["A" + str(i)].value == str(message.author.id):
                #sheet["B" + str(i)].balue = sheet["B" + str(i)].balue + 1
                #if sheet ["B" + str(i)].balue >= exp[sheet["C" + str(i)].value - 1]:
                    #sheet["C" + str(i)].value = sheet["C" + str(i)].value + 1e
                    #await message.channel.send("레벨이 올랐습니다. \n현재 레벨 : " + str(sheet["C" + str(i)].value) + "\n경험치 : " + str(sheet["B" + str(i)].value))         
                    #file.save("레벨.xlsx")
                #break
            #if sheet["A" + str(i)].value == None:
                #sheet["A" + str(i)].value = str(message.author.id)
                #sheet["B" + str(i)].value = 0
                #sheet["C" + str(i)].value = 0
                #file.save("레벨.xlsx")
                #break
            #i += 1
    

    #if message.content == ("진봇 음성채널 입장"):
        #await message.author.voice.channel.connect()
        #await message.channel.send("보이스 채널에 입장합니다")


    #if message.content == ("진봇 음성채널 퇴장"):
        #for vc in client.voice_clients:
            #if vc.guild == message.guild:
                #voice = vc
        #await vc.disconnect()
        #await message.channel.send("보이스 채널에서 퇴장합니다")

    #if message.content.startswith("진봇 URL재생"):
    #   for vc in client.voice_clients:
    #      if vc.guild == message.guild:
    #         voice = vc
        #url = message.content.split(" ")[1]
        #option = {
        #   "outtmpl" : "file/" + url.split("=")[1] + ".mp3"
        #}
        

    #with youtube_dl.YoutubeDL(option) as ydl:
    #   ydl.download([url])
    #  info = ydl.extract_info(url, download=False)
    # title = info["title"]
        #voice.play(discord.FFmpegPCMAudio("file/" + url.split("=")[1] + ".mp3"))
        #await message.channel.send(title + "을 재생합니다")

    if message.content == "진봇 역사퀴즈":
        quiz = int(random() * 12)
        if quiz == 1:
            await message.channel.send("청동기 시대에 커다란 돌을 쌓아 만든 족장의 무덤은?||고인돌||")
        if quiz == 2:
            await message.channel.send("우리 나라 최초의 국가는?\n||고조선||")
        if quiz == 3:
            await message.channel.send("'널리 인간을 이롭게 한다.'는 고조선의 건국 이념은?\n||홍익인간||")
        if quiz == 4:
            await message.channel.send("단군 왕검이 하늘에 제사를 지냈다는 참성단이 있는 곳은?\n||강화도||")
        if quiz == 5:
            await message.channel.send("동명 성왕인 주몽이 세운 나라의 이름은?\n||고구려||")
        if quiz == 6:
            await message.channel.send("삼국이 통일된 후 대조영이 옛 고구려의 영토에 세운 나라는?\n||발해||")
        if quiz == 7:
            await message.channel.send("신라가 삼국 통일의 위업을 달성할 때의 왕은 누구인가?\n||문무왕||")
        if quiz == 8:
            await message.channel.send("화랑도 정신인 세속 오계를 지은 사람은?\n||원광 법사||")
        if quiz == 9:
            await message.channel.send("백제의 영토를 북쪽으로는 한강, 남쪽으로는 탐라국까지 넓힌 왕은 누구인가?\n||근초고왕||")
        if quiz == 10:
            await message.channel.send("고구려의 영토를 북만주까지 넓혀 고구려인의 용맹을 과시한 왕은 누구인가?\n||광개토대왕||")
        if quiz == 11:
            await message.channel.send("고려에 쳐들어온 외적으로 만주 서북부 지방에 살았고 요라는 나라를 세운 민족은?\n ||거란족||")
        if quiz == 12:
            await message.channel.send("만주 동북부 지방에 살면서 고려의 국경을 자주 넘어와서 노략질을 하였으면 금이라는 나라를 세운 민족은?\n||여진족||")
        if quiz == 13:
            await message.channel.send("고려와 친교를 맺고 사이좋게 지냈던 나라는?\n||송나라||")
    if message.content.startswith("진봇 정보"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        embed = discord.Embed(color=0x00ff00)
        embed.add_field(name="이름", value=message.author.name, inline=False)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=False)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일", inline=False)
        embed.add_field(name="아이디", value=message.author.id, inline=False)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(message.channel, embed=embed)

    if message.content.startswith("진봇 투표"):
        vote = message.content[6:].split("/")
        await message.channel.send(message.channel, vote[0])
        for i in range(1, len(vote)):
            choose = await message.channel.send(message.channel, vote[i])
            await message.add_reation(choose, "👍")

    if message.content.startswith("진봇 서버목록"):
        list = []
        for guild in client.guilds:
            list.append(guild.name)
        await message.channel.send("\n".join(list))

    if message.content.startswith("진봇 현재시각"):
        a = datetime.datetime.today().year
        b = datetime.datetime.today().month
        c = datetime.datetime.today().day
        d = datetime.datetime.today().hour
        e = datetime.datetime.today().minute
        f = datetime.datetime.today().second
        await message.channel.send(str(a) + "년" + str(b) + "월" + str(c) + "일" + str(d) + "시" + str(e) + "분" + str(f) + "초 입니다.")

    if message.content.startswith("진봇 팀정하기"):
        team = message.content[8:]
        peopleteam = team.split("/")
        people = peopleteam[0]
        team = peopleteam[1]
        person = people.split(" ")
        teamname = team.split(" ")
        random.shuffle(teamname) 
        for i in range(0, len(person)):
            await message.channel.send(person[i] + "------>" + teamname[i])

    if message.content.startswith("진봇 롤"):
        name = message.content[5:]
        req = requests.get("http://www.op.gg/summoner/userName=" + name)
        html = req.text
        soup = BeautifulSoup(html, "html.parser")

        Rank1 = soup.find_all("span", {"class": "tierRank"})
        Rank2 = str(Rank1[0])[str(Rank1[0]).find('nk">') + 4:str(Rank1[0]).find("</span")]

        lp1 = soup.find_all("span", {"class": "LeaguePoints"})
        lp2 = str(lp1[0])[str(lp1[0]).find('">')+2:str(lp1[0]).find(">/sp")]
        lp3 = lp2.strip()

        win1 = soup.find_all("span", {"class": "WinLose"})
        win2 = str(win1[0])[str(win1[0]).find('ns">')+2:str(win1[0]).find(">/sp")]
        win3 = win2.replace("W", "승")

        lose1 = soup.find_all("span", {"class": "losses"})
        lose2 = str(lose1[0])[str(lose1[0]).find('es">')+2:str(lose1[0]).find(">/sp")]
        lose3 = lose2.replace("L", "패")
        

        ratio1 = soup.find_all("span", {"class": "winratio"})
        ratio2 = str(ratio1[0])[str(ratio1[0]).find('io">')+2:str(ratio1[0]).find(">/sp")]
        ratio3 = ratio2.replace("Win Ratio", "승률")

        embed = discord.Embed(color=0x00ff00)
        embed.add_field(name="랭크", value=Rank2, inline=True)
        embed.add_field(name="레벨", value=lp3, inline=True)
        embed.add_field(name="승리 수", value=win3, inline=True)
        embed.add_field(name="패배 수", value=lose3, inline=True)
        embed.add_field(name="승률", value=ratio3, inline=True)
        await message.channel.send(embed=embed)




    if message.content.startswith("진봇 아"):
        await message.channel.send("아아아ㅏ아아아ㅏ아아아")
    if message.content.startswith("진봇 ㅇㅉㅌㅂ"):
        await message.channel.send("어쩔티비 저쩔티비 응 빡치쥬~? 응 화나쥬? 응 아무것도 못하쥬? ")
    if message.content.startswith("진봇 저쩔티비"):
        await message.channel.send("ㅇㅉ")
    if message.content.startswith("진봇 어쩔냉장고"):
        await message.channel.send("안물안궁")
    if message.content.startswith("진봇 바보"):
        await message.channel.send("자기소개 잘하누 ㅋㅋ")
    if message.content.startswith("진봇 ?"):
        await message.channel.send("궁금하면 도움말 보시등가 ㅋㅋ")
    if message.content.startswith("진봇 모여봐요 동물의숲"):
        await message.channel.send("갚아봐요 사채의숲")
    if message.content.startswith("진봇 샌즈"):
        await message.channel.send("와! 샍으 아시는군아!")
    if message.content == "진봇 ㄷ":
        await message.channel.send("역시 나란... 사람들이 다 무서워하는군...")
    if message.content == "진봇 종료":
        await message.channel.send("진봇을 종료합니다!(진봇의 개발자가 아닐시 메시지만 나오고 실제로 종료되진 않습니다)")
    #if message.content == "진봇 도배":
    if message.content == ("진봇 자폭"):
        await message.channel.send("무지개반사 너나자폭해")
    if message.content == ("진봇 죽어"):
        await message.channel.send("나 데이터쪼가리라서 못죽는다니깐?")
    if message.content == "진봇 ㅇㅁㅇ":
        await message.channel.send("ㄴㅇㄱ")
    if message.content == "진봇 꺼져":
        await message.channel.send("ㅇㅉㅌㅂ")
    if message.content.startswith("진봇 따라해"):
        moge = message.content[7:].split(" ")
        await message.channel.send("ㅇㅇ " + str(moge))
    if message.content == "진봇 망할놈":
        await message.channel.send("데이터쪼가리를 싫어해서 어쩌려고?")
    if message.content == "진봇 초대링크":
        await message.channel.send("https://discord.com/api/oauth2/authorize?client_id=928874680226566174&permissions=8&scope=bot")
    if message.content == "진봇 개발자 좋아해?":
        await message.channel.send("ㄴ")
    if message.content == "진봇 진 좋아해?":
        await message.channel.send("ㄴ")
    if message.content == "진봇 진봇 좋아해?":
        await message.channel.send("ㄴ")
    if message.content == "진봇 개인정보 얼마야?":
        await message.channel.send("그걸 물어보기 전에 니 개인정보부터 내놔 ㅋㅎ 그럼 내가 니 개인정보 가격 알아올게")
    if message.content == "진봇 ㅗ": 
        await message.channel.send("한글 자모의 열아홉 번째 글자이자 모음의 다섯 번째 글자. '오'라고 읽는다.\n로마자로 쓸때 주로 O라고 쓰고, 국제음성기호로 단모음은 후설 원순 중고모음 [o]로 나타낸다. 이 발음은 ㄱ계 자음과 ㅎ을 구개수음화 시키는 성질이 있는데, 이 구개수음화를 피하기 위해 [o̞]로 발음할 수도 있다. 훈민정음에서는 ㆍ와 ㅡ를 합해 만들었고, 그 소리를 'ㆍ와 같지만 입을 오므린다'고 설명하고 있다. 중성 모음 ㅣ와 합쳐져 단모음 ㅚ을, 양성 모음 ㅏ와 합쳐져 이중 모음 ㅘ를, 그리고 ㅐ와 합쳐져 이중 모음 ㅙ를 표기할 수 있다. 훈몽자회에서는 吾와 음가가 같다고 설명하고 있다./나무위키")
    if message.content == "진봇 짖어":
        await message.channel.send("내가왜 니나짖어")
    if message.content == "진봇 멍멍":
        await message.channel.send("아이고 잘한다~")
    if message.content == "진봇 뒤져":
        await message.channel.send("아쉽게도 봇이라 데이터쪼가리여서 뒤질수가 없어요! 죄송합니다!")
    if message.content == "진봇 신기해":
        await message.channel.send("~~이게 왜신기함?~~")
    if message.content == "진봇 어쩔티비":
        await message.channel.send("어쩔티비저쩔티비안물티비않궁티비")
    if message.content == "진봇 ㅋㅋㄹㅃㅃ":
        await message.channel.send("ㅇㅉㅌㅂ")
    if message.content == "진봇 마크":
        await message.channel.send("||~~망겜~~||")
    if message.content == "진봇 낮추고 지키고 움직이지 않기":
        await message.channel.send("ERROR! 그것을 실행하기엔 니놈은 너무 하찮아!")
    if message.content == "진봇 임베드":
        embed = discord.Embed(title="임베드", description="엄청나", color=0x00ff00)
        await message.channel.send(embed=embed)
    if message.content.startswith("진봇 아재"):
        await message.channel.send("아니다")
    if message.content.startswith("진봇 하픽"):
        await message.channel.send("세계 최강의 망겜 서버!")
    if message.content.startswith("진봇 지금우리진봇은"):
        await message.channel.send("천재")
    if message.content == "진봇 샌즈":
        await message.channel.send("는 샌즈다")
    if message.content == "진봇 나는야":
        await message.channel.send("멍청이(너가)")
    if message.content == "진봇 멍청이":
        await message.channel.send("니가")
    if message.content == "진봇 강기율":
        await message.channel.send("멍청이 이거 기율한테 비밀로 좀")
    
    if message.content == "진봇 서버정보":
        severicon = f"{message.guild.icon}"       
        await message.channel.send(f"서버이름:{message.guild.name}\n서버아이디:{message.guild.id}\n실행한 사람:{message.author.mention}\n서버 아이콘: https://cdn.discordapp.com/icons/{message.guild.id}/{message.guild.icon}.webp?size=96")

        if message.content.startswith("진봇 기억"):
            file = openpyxl.load_workbook("기억.xlsx")
            sheet = file.active
            memory = message.content.split("/")
            for i in range(1, 59999999999999999999999999999999999):
                if sheet["A" + str(i)].value == memory[1]:
                    await message.channel.send(sheet["B" + str(i)].value)
                    break

    if message.content == "뿗뻵쀨뿕":
        while True: 
            await message.channel.send("@Jin_3070#9999")

bot.command()



access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
